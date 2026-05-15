from pathlib import Path
import unicodedata

from openpyxl import load_workbook


ARQUIVO_EXCEL = "base pesquisa 5 estrelas março 26.xlsx"
COLUNA_ALVO = "NOTA GERAL"


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace(" ", "").replace("_", "")
    return text


def to_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if text == "":
        return None

    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def main() -> None:
    path = Path(ARQUIVO_EXCEL)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")

    target = normalize_header(COLUNA_ALVO)
    soma = 0.0
    quantidade = 0
    abas_com_coluna = 0

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                continue

            idx_coluna = None
            for idx, col_name in enumerate(header):
                if normalize_header(col_name) == target:
                    idx_coluna = idx
                    break

            if idx_coluna is None:
                continue

            abas_com_coluna += 1
            for row in rows:
                if idx_coluna >= len(row):
                    continue
                numero = to_number(row[idx_coluna])
                if numero is None:
                    continue
                soma += numero
                quantidade += 1
    finally:
        workbook.close()

    if quantidade == 0:
        print("Nenhum valor numérico válido encontrado na coluna NOTA GERAL.")
        return

    media = soma / quantidade
    print(f"Arquivo: {ARQUIVO_EXCEL}")
    print(f"Abas com coluna NOTA GERAL: {abas_com_coluna}")
    print(f"Quantidade de valores: {quantidade}")
    print(f"Média de NOTA GERAL: {media:.4f}")


if __name__ == "__main__":
    main()
